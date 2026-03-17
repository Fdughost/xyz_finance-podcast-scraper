#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel报表生成模块 - 优化版
将爬取的播客数据导出为Excel格式，确保所有字段完整展示
"""

import pandas as pd
import os
from datetime import datetime
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class PodcastExcelGenerator:
    """播客Excel报表生成类"""
    
    def __init__(self, output_dir: str = 'reports'):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_daily_report(self, data: list, filename: str = None) -> str:
        """生成每日Excel报表"""
        if not data:
            logger.warning("没有数据，无法生成报表")
            return ""

        today = datetime.now().strftime('%Y-%m-%d')
        if filename is None:
            filename = f'播客监控日报_{today}.xlsx'
        
        file_path = os.path.join(self.output_dir, filename)

        # 转换为DataFrame
        df = pd.DataFrame(data)
        
        # 字段映射（中文名称）
        column_mapping = {
            'title': '节目名称',
            'category': '分类',
            'institution_name': '机构名称',
            'subscribers': '订阅数',
            'latest_episode_title': '最新单集名称',
            'latest_episode_date': '最新单集上线日期',
            'latest_episode_play': '最新单集播放数量',
            'latest_episode_comment': '最新单集评论数',
            'latest_episode_clap': '最新单集点赞数',
            'latest_episode_favorite': '最新单集收藏数',
            'latest_episode_interaction': '互动指标(点赞+收藏)',
            'latest_episode_duration': '最新单集时长',
            'crawl_time': '抓取时间'
        }
        
        # 确保所有列都存在，不存在的填充为空字符串
        for col in column_mapping.keys():
            if col not in df.columns:
                df[col] = ""
        
        # 按订阅数降序排列
        df['subscribers'] = pd.to_numeric(df['subscribers'], errors='coerce').fillna(0).astype(int)
        df = df.sort_values(by='subscribers', ascending=False).reset_index(drop=True)

        # 按照映射顺序重新排序列并重命名
        df = df[list(column_mapping.keys())].rename(columns=column_mapping)

        # 写入Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='每日监控数据')
            
            workbook = writer.book
            worksheet = writer.sheets['每日监控数据']
            
            # 设置样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

            # 应用样式到表头
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
                
                # 设置初始列宽
                column_letter = get_column_letter(col_num)
                if column_title in ['最新单集名称']:
                    worksheet.column_dimensions[column_letter].width = 40
                elif column_title in ['互动指标(点赞+收藏)']:
                    worksheet.column_dimensions[column_letter].width = 22
                elif column_title in ['最新单集上线日期', '抓取时间']:
                    worksheet.column_dimensions[column_letter].width = 20
                else:
                    worksheet.column_dimensions[column_letter].width = 15

            # 应用样式到数据行
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
                for cell in row:
                    cell.alignment = left_alignment
                    cell.border = border
                    # 为数字列设置居中
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = center_alignment

        logger.info(f"Excel报表已生成: {file_path}")
        return file_path
